from __future__ import annotations

import re
from pathlib import Path
from typing import Optional, List

import openpyxl

from .template_model import SubjectTemplate, TemplateItem, Block


_RE_NUM = re.compile(r"(\d+(?:[.,]\d+)?)")

def _parse_max_points(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        nums = _RE_NUM.findall(v)
        if not nums:
            return 0.0
        return float(sum(float(x.replace(",", ".")) for x in nums))
    return 0.0

def _parse_week(v) -> Optional[int]:
    if v is None:
        return None
    # В некоторых файлах первая колонка может быть datetime или мусор
    if isinstance(v, (int, float)):
        w = int(v)
        if 1 <= w <= 15:
            return w
        return None
    return None

def _detect_block(cell_value) -> Optional[Block]:
    if not isinstance(cell_value, str):
        return None
    s = cell_value.strip().lower()
    if "рейтинг 1" in s or "rk1" in s or "рк1" in s:
        return "RK1"
    if "рейтинг 2" in s or "rk2" in s or "рк2" in s:
        return "RK2"
    return None

def parse_subject_template(xlsx_path: str | Path, subject_id: str, subject_name: str) -> SubjectTemplate:
    path = Path(xlsx_path)
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb[wb.sheetnames[0]]

    items: List[TemplateItem] = []
    current_block: Optional[Block] = None

    # Ожидаем заголовок на первой строке, данные далее
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value  # Week / block label
        b = ws.cell(r, 2).value  # type
        c = ws.cell(r, 3).value  # title
        d = ws.cell(r, 4).value  # max points

        block = _detect_block(a)
        if block:
            current_block = block
            continue

        if current_block is None:
            continue

        week = _parse_week(a)
        max_points = _parse_max_points(d)

        # Пропускаем пустые/нулевые строки и строки "ИТОГО/ЖИЫНЫ ..."
        if week is None:
            continue
        if max_points <= 0:
            continue
        # title может быть None
        title = (str(c).strip() if c is not None else (str(b).strip() if b is not None else ""))
        if not title:
            title = "Item"

        items.append(
            TemplateItem(
                subject_id=subject_id,
                subject_name=subject_name,
                block=current_block,
                week=week,
                title=title,
                max_points=float(max_points),
            )
        )

    return SubjectTemplate(subject_id=subject_id, subject_name=subject_name, items=items)